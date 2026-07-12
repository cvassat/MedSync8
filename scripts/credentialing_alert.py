#!/usr/bin/env python3
"""NEH Credentialing — Daily Risk Alert.

Reads the star-schema workbook (FactLicensing + DimEmployee + DimCredential
+ DimStatus), identifies every credential expiring within 90 days, and posts
a ranked Teams MessageCard.

Schedule:  daily 08:00 America/Chicago
Owner:     Codie Vassar, MD · NEH Psychiatry
PHI:       none — aggregate risk metadata only
"""
from __future__ import annotations

import json
import logging
import os
import sys
import urllib.request
from dataclasses import dataclass
from datetime import date, datetime
from typing import Any

import openpyxl

log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Configuration — override via environment variables in production
# ---------------------------------------------------------------------------

WORKBOOK_PATH = os.environ.get(
    "CREDENTIALING_WORKBOOK",
    "/home/user/workspace/2026-06-28_Template_Licensing-Credentialing-StarSchema_v1_draft.xlsx",
)

# Teams incoming-webhook URL (required).  Set in the scheduler / GitHub
# Actions secret.  A second URL (e.g. a #credentialing channel) is optional.
TEAMS_WEBHOOK_URL    = os.environ.get("TEAMS_WEBHOOK_URL", "")
TEAMS_CHANNEL_WEBHOOK = os.environ.get("TEAMS_CHANNEL_WEBHOOK", "")

CRITICAL_DAYS = 30
URGENT_DAYS   = 60
WATCH_DAYS    = 90

# Module-level constant so rank_key() doesn't rebuild it on every call.
_BAND_ORDER = {
    "LAPSED": 0, "CRITICAL": 1, "URGENT": 2,
    "WATCH": 3, "STALLED": 4, "PENDING": 5, "OK": 6,
}

BAND_EMOJI = {
    "LAPSED":   "🛑",
    "CRITICAL": "🔴",
    "URGENT":   "🟠",
    "WATCH":    "🟡",
    "STALLED":  "⚪",
    "PENDING":  "⚪",
}

# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------

@dataclass
class LicenseRow:
    license_id: int
    employee_id: int
    credential_id: int
    status_id: int
    issue_date: date | None
    expiry_date: date | None
    notes: str = ""
    employee_name: str = ""
    employee_role: str = ""
    supervising_id: int | None = None
    credential_name: str = ""
    credential_type: str = ""
    jurisdiction: str = ""
    issuing_authority: str = ""
    status_name: str = ""
    status_category: str = ""
    risk_weight: int = 0
    days_to_expiry: int | None = None
    urgency_band: str = ""
    action_required: str = ""

# ---------------------------------------------------------------------------
# Workbook loading
# ---------------------------------------------------------------------------

def _rows_from_table(ws: openpyxl.worksheet.worksheet.Worksheet,
                     table_name: str) -> list[dict[str, Any]]:
    """Return rows of an Excel Table as column-keyed dicts.

    Blank header cells get a fallback name (_col0, _col1 …) so they never
    collide under the None key.
    """
    if table_name not in ws.tables:
        raise KeyError(f"Table {table_name!r} not found on sheet {ws.title!r}")
    cells = list(ws[ws.tables[table_name].ref])
    # FIX #8: replace None headers so multiple blank columns don't collide
    header = [
        (c.value if c.value is not None else f"_col{i}")
        for i, c in enumerate(cells[0])
    ]
    out: list[dict[str, Any]] = []
    for row in cells[1:]:
        vals = [c.value for c in row]
        if all(v is None for v in vals):
            continue
        out.append(dict(zip(header, vals)))
    return out


def _coerce_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    # FIX #1: handle dates stored as ISO strings in Excel
    if isinstance(value, str):
        try:
            return date.fromisoformat(value[:10])
        except ValueError:
            return None
    return None


def _load_dim(wb: openpyxl.Workbook, sheet: str, table: str,
              key_col: str) -> dict[Any, dict[str, Any]]:
    """Load a dimension table; warn on duplicate keys (FIX #5)."""
    rows = _rows_from_table(wb[sheet], table)
    out: dict[Any, dict[str, Any]] = {}
    for r in rows:
        k = r[key_col]
        if k in out:
            log.warning("Duplicate %s=%s in %s — later row wins", key_col, k, table)
        out[k] = r
    return out


def load_workbook_data(path: str, today: date) -> list[LicenseRow]:
    wb = openpyxl.load_workbook(path, data_only=True)
    try:  # FIX #9: ensure handle is closed even on error
        fact_rows = _rows_from_table(wb["FactLicensing"], "FactLicensing")
        dim_emp  = _load_dim(wb, "DimEmployee",   "DimEmployee",   "EmployeeID")
        dim_cred = _load_dim(wb, "DimCredential", "DimCredential", "CredentialID")
        dim_stat = _load_dim(wb, "DimStatus",     "DimStatus",     "StatusID")

        enriched: list[LicenseRow] = []
        for r in fact_rows:
            # FIX #3: catch per-row KeyError so one bad row doesn't abort the run
            try:
                lic = LicenseRow(
                    license_id    = r["LicenseID"],
                    employee_id   = r["EmployeeID"],
                    credential_id = r["CredentialID"],
                    status_id     = r["StatusID"],
                    issue_date    = _coerce_date(r.get("IssueDate")),
                    expiry_date   = _coerce_date(r.get("ExpiryDate")),
                    notes         = (r.get("Notes") or "").strip(),
                )
            except KeyError as exc:
                log.warning("Skipping FactLicensing row — missing column %s: %r", exc, r)
                continue

            emp = dim_emp.get(lic.employee_id, {})
            lic.employee_name  = emp.get("FullName", f"Employee #{lic.employee_id}")
            lic.employee_role  = emp.get("Role", "")
            # FIX: coerce empty string to None (supervising_id typed int | None)
            lic.supervising_id = emp.get("SupervisingPhysicianID") or None

            cred = dim_cred.get(lic.credential_id, {})
            lic.credential_name   = cred.get("CredentialName", f"Credential #{lic.credential_id}")
            lic.credential_type   = cred.get("CredentialType", "")
            lic.jurisdiction      = cred.get("Jurisdiction", "")
            lic.issuing_authority = cred.get("IssuingAuthority", "")

            stat = dim_stat.get(lic.status_id, {})
            lic.status_name     = stat.get("StatusName", "")
            lic.status_category = stat.get("StatusCategory", "")
            lic.risk_weight     = int(stat.get("RiskWeight") or 0)

            if lic.expiry_date is not None:
                lic.days_to_expiry = (lic.expiry_date - today).days

            enriched.append(lic)

        return enriched
    finally:
        wb.close()

# ---------------------------------------------------------------------------
# Risk classification
# ---------------------------------------------------------------------------

def classify(lic: LicenseRow, today: date) -> None:  # FIX #4: accept today
    """Assign urgency_band and action_required.  Mutates lic in place."""
    d = lic.days_to_expiry

    if d is None:
        # FIX #4: use the passed today, not date.today()
        if lic.issue_date and (today - lic.issue_date).days > 60:
            lic.urgency_band = "STALLED"
            lic.action_required = "Follow up with issuing authority — application open >60 days"
        else:
            lic.urgency_band = "PENDING"
            lic.action_required = "Monitor issuance"
        return

    if d < 0:
        lic.urgency_band = "LAPSED"
        lic.action_required = "STOP-WORK — credential expired. Do not schedule; verify scope."
    elif d <= CRITICAL_DAYS:
        lic.urgency_band = "CRITICAL"
        lic.action_required = "Renewal must be submitted THIS WEEK. Notify supervisor."
    elif d <= URGENT_DAYS:
        lic.urgency_band = "URGENT"
        lic.action_required = "Cross the 60-day cliff — open renewal file now."
    elif d <= WATCH_DAYS:
        lic.urgency_band = "WATCH"
        lic.action_required = "Renewal window open — schedule application."
    else:
        # FIX #10: explicit OK so urgency_band is never left as ""
        lic.urgency_band = "OK"
        lic.action_required = ""
        return  # no cross-check needed for out-of-window credentials

    # Data-integrity cross-check
    computed_expiring = 0 <= d <= WATCH_DAYS
    # FIX #7: warn on any non-expiring status, not just "Active"
    expiring_statuses = {"Expiring Soon", "Expired", "Suspended", "Inactive"}
    if computed_expiring and lic.status_name not in expiring_statuses:
        lic.action_required += f"  ⚠ DimStatus shows '{lic.status_name}' — update StatusID."


def rank_key(lic: LicenseRow) -> tuple[int, int, int]:
    d = lic.days_to_expiry if lic.days_to_expiry is not None else 10_000
    return (_BAND_ORDER.get(lic.urgency_band, 9), d, -lic.risk_weight)

# ---------------------------------------------------------------------------
# Teams delivery
# ---------------------------------------------------------------------------

def _build_teams_card(rows: list[LicenseRow], today: date) -> dict[str, Any]:
    """Compose an Office-365-connector MessageCard payload."""
    actionable = sorted(
        [r for r in rows if r.urgency_band in {"LAPSED", "CRITICAL", "URGENT", "WATCH", "STALLED"}],
        key=rank_key,
    )

    if not actionable:
        return {
            "@type": "MessageCard",
            "@context": "https://schema.org/extensions",
            "themeColor": "00C176",
            "summary": f"NEH Credentialing Alert · {today.isoformat()}",
            "title": f"NEH Credentialing — Daily Alert · {today.isoformat()}",
            "text": "✅ No credentials in the 90-day risk window. Portfolio is compliant.",
        }

    counts: dict[str, int] = {}
    for r in actionable:
        counts[r.urgency_band] = counts.get(r.urgency_band, 0) + 1

    summary_parts = [
        f"{BAND_EMOJI[b]} {counts[b]} {b.title()}"
        for b in ("LAPSED", "CRITICAL", "URGENT", "WATCH", "STALLED")
        if counts.get(b)
    ]
    theme = "FF4444" if (counts.get("LAPSED") or counts.get("CRITICAL")) else "FF8C00"

    # Cap at 20 entries so the payload stays under Teams' 28 KB limit.
    display = actionable[:20]
    sections = []
    for i, r in enumerate(display, start=1):
        expiry_str = r.expiry_date.isoformat() if r.expiry_date else "—"
        d = r.days_to_expiry
        days_str = (
            f"expired {-d}d ago" if d is not None and d < 0
            else (f"{d}d" if d is not None else "no expiry set")
        )
        emoji = BAND_EMOJI.get(r.urgency_band, "•")
        sections.append({
            "activityTitle": (
                f"{emoji} {i}. {r.employee_name} — {r.credential_name} ({r.jurisdiction})"
            ),
            "activityText": (
                f"Expires **{expiry_str}** ({days_str})"
                f" · Status: `{r.status_name}` · Risk: {r.risk_weight}<br>"
                f"→ {r.action_required}"
            ),
        })

    overflow_note = (
        f"\n\n_{len(actionable) - 20} additional items not shown — review the full dashboard._"
        if len(actionable) > 20 else ""
    )

    return {
        "@type": "MessageCard",
        "@context": "https://schema.org/extensions",
        "themeColor": theme,
        "summary": (
            f"NEH Credentialing Alert · {today.isoformat()} · {len(actionable)} item(s)"
        ),
        "title": f"NEH Credentialing — Daily Alert · {today.isoformat()}",
        "text": "  ·  ".join(summary_parts) + overflow_note,
        "sections": sections,
    }


def send_teams(rows: list[LicenseRow], today: date) -> bool:
    """POST a MessageCard to every configured Teams webhook.

    Returns True only if every delivery succeeded.  FIX #2: uses a 15-second
    timeout so a stalled webhook never blocks the process indefinitely.
    FIX #6: returns False on any failure so main() can exit non-zero.
    """
    if not TEAMS_WEBHOOK_URL:
        log.error("TEAMS_WEBHOOK_URL is not set — alert not delivered")
        return False

    card = _build_teams_card(rows, today)
    payload = json.dumps(card, ensure_ascii=False).encode("utf-8")

    targets = [TEAMS_WEBHOOK_URL]
    if TEAMS_CHANNEL_WEBHOOK:
        targets.append(TEAMS_CHANNEL_WEBHOOK)

    all_ok = True
    for url in targets:
        req = urllib.request.Request(
            url,
            data=payload,
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        try:
            with urllib.request.urlopen(req, timeout=15) as resp:  # FIX #2: timeout
                if resp.status not in (200, 202):
                    log.warning("Teams webhook returned HTTP %s", resp.status)
                    all_ok = False
                else:
                    log.info("Teams alert delivered → %s…", url[:60])
        except Exception as exc:
            log.warning("Teams delivery failed for %s…: %s", url[:60], exc)
            all_ok = False

    return all_ok

# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> int:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)-8s %(message)s",
        datefmt="%Y-%m-%dT%H:%M:%S",
    )

    today = date.today()

    # FIX #3: catch workbook-structure errors, not just FileNotFoundError
    try:
        rows = load_workbook_data(WORKBOOK_PATH, today)
    except FileNotFoundError:
        log.error("Workbook not found: %s", WORKBOOK_PATH)
        return 2
    except (KeyError, ValueError,
            openpyxl.utils.exceptions.InvalidFileException) as exc:
        log.error("Failed to load workbook: %s", exc)
        return 2

    for r in rows:
        classify(r, today)  # FIX #4: pass today

    actionable = [r for r in rows if r.urgency_band not in {"OK", "PENDING", ""}]
    actionable.sort(key=rank_key)
    log.info("Loaded %d credential rows; %d require attention", len(rows), len(actionable))
    for r in actionable:
        log.info("  [%-8s] %s — %s (%s)", r.urgency_band,
                 r.employee_name, r.credential_name, r.jurisdiction)

    # FIX #6: propagate delivery failure to exit code so the scheduler alerts
    ok = send_teams(rows, today)
    return 0 if ok else 1


if __name__ == "__main__":
    raise SystemExit(main())
